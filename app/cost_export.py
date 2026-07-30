"""Build Excel (.xlsx) and PDF exports for cost calculator results."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _money(n: Any) -> str:
    try:
        return f"${float(n or 0):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def _site_lines(meta: dict[str, Any] | None) -> list[str]:
    if not meta:
        return []
    parts = []
    if meta.get("road_name"):
        parts.append(str(meta["road_name"]))
    if meta.get("site_number"):
        parts.append(f"Site {meta['site_number']}")
    if meta.get("moa_number"):
        parts.append(f"MoA {meta['moa_number']}")
    if meta.get("tgs_reference"):
        parts.append(f"TGS {meta['tgs_reference']}")
    return parts


def _booking_rows(result: dict[str, Any]) -> list[str]:
    items = result.get("booking_requirements") or []
    if items:
        return [b.get("text") or f"{b.get('quantity')}× {b.get('label')}" for b in items]
    summary = result.get("booking_summary") or ""
    return [summary] if summary else []


def build_cost_workbook(
    result: dict[str, Any],
    *,
    title: str = "Cost estimate",
    site: dict[str, Any] | None = None,
    notes: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    best_fill = PatternFill("solid", fgColor="C8E6C9")
    money_font = Font(bold=True)

    row = 1
    ws.cell(row, 1, "WRU TGS Tracker — Traffic cost export").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row, 1, title)
    row += 1
    ws.cell(row, 1, f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 1
    for line in _site_lines(site):
        ws.cell(row, 1, line)
        row += 1
    if notes:
        ws.cell(row, 1, f"Notes: {notes}")
        row += 1
    row += 1

    mode = result.get("mode")
    ws.cell(row, 1, "Mode").font = Font(bold=True)
    ws.cell(row, 2, mode or "")
    row += 2

    booking = _booking_rows(result)
    if booking:
        ws.cell(row, 1, "Booking requirements").font = Font(bold=True)
        row += 1
        for b in booking:
            ws.cell(row, 1, b)
            row += 1
        row += 1

    if mode == "standard":
        echo = result.get("inputs_echo") or {}
        per = result.get("per_shift") or {}
        allow = per.get("allowances") or {}
        rows = [
            ("Shift type", echo.get("shift_type")),
            ("Shift start time", echo.get("shift_start_time")),
            ("Shift hours", echo.get("shift_hours")),
            ("Days of work", echo.get("days_of_work")),
            ("Shifts per day", echo.get("shifts_per_day")),
            ("Total shifts", echo.get("total_shifts")),
            ("Works start", echo.get("works_start")),
            ("Works end", echo.get("works_end")),
            ("Per-shift pack labour", _money(per.get("shift_labour_total"))),
            ("Per-shift travel", _money(allow.get("travel_total"))),
            ("Per-shift meals", _money(allow.get("meal_total"))),
            ("Per-shift total", _money(per.get("shift_total"))),
            ("Site crew total", _money(result.get("site_crew_total"))),
            ("Site allowances", _money(result.get("site_allowances_total"))),
            ("VMS total", _money((result.get("vms") or {}).get("vms_total"))),
            ("Site traffic total", _money(result.get("site_traffic_total"))),
        ]
        for label, val in rows:
            ws.cell(row, 1, label)
            ws.cell(row, 2, val)
            row += 1

        # Pack lines sheet
        ws2 = wb.create_sheet("Pack mix")
        headers = [
            "Qty",
            "Booking",
            "Kind",
            "TCs",
            "Vehicles",
            "Ord h",
            "OT h",
            "Ord $",
            "OT $",
            "Line total",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
        for i, line in enumerate(per.get("lines") or [], start=2):
            ws2.cell(i, 1, line.get("quantity"))
            ws2.cell(i, 2, line.get("booking_label") or line.get("name"))
            ws2.cell(i, 3, line.get("rate_kind"))
            ws2.cell(i, 4, line.get("people_covered"))
            ws2.cell(i, 5, line.get("vehicles_covered"))
            ws2.cell(i, 6, line.get("ordinary_hours"))
            ws2.cell(i, 7, line.get("overtime_hours"))
            ws2.cell(i, 8, line.get("ordinary_rate"))
            ws2.cell(i, 9, line.get("overtime_rate"))
            ws2.cell(i, 10, line.get("line_total"))

    elif mode == "closure_24h":
        rec = result.get("recommendation") or {}
        ws.cell(row, 1, "Recommendation").font = Font(bold=True)
        ws.cell(row, 2, rec.get("best_label") or rec.get("cheaper"))
        row += 1
        ws.cell(row, 1, "Summary")
        ws.cell(row, 2, rec.get("summary"))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 2

        headers = [
            "Pattern",
            "Shift hours",
            "Shifts",
            "Pack labour",
            "Travel",
            "Meals",
            "Crew total",
            "VMS",
            "Grand total",
            "Best?",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row, col, h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for key in ("option_3x8", "option_2x12"):
            opt = result.get(key) or {}
            vals = [
                opt.get("label"),
                opt.get("shift_hours"),
                opt.get("shifts_required"),
                opt.get("pack_labour_total"),
                opt.get("travel_total"),
                opt.get("meal_total"),
                opt.get("crew_total") or opt.get("labour_total"),
                opt.get("vms_total"),
                opt.get("grand_total"),
                "BEST" if opt.get("is_best") else "",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row, col, val)
                if opt.get("is_best"):
                    cell.fill = best_fill
                    if col == 9:
                        cell.font = money_font
            row += 1

        # Detail sheets per option
        for key, sheet_name in (("option_3x8", "8h shifts"), ("option_2x12", "12h shifts")):
            opt = result.get(key) or {}
            ws_d = wb.create_sheet(sheet_name[:31])
            ws_d.cell(1, 1, opt.get("label") or sheet_name).font = Font(bold=True)
            ws_d.cell(2, 1, f"Best option: {'YES' if opt.get('is_best') else 'no'}")
            ws_d.cell(3, 1, "Booking")
            ws_d.cell(3, 2, opt.get("booking_summary") or "")
            headers = [
                "#",
                "Type",
                "Hours",
                "Pack labour",
                "Travel",
                "Meals",
                "Shift total",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws_d.cell(5, col, h)
                cell.font = header_font
                cell.fill = header_fill
            for i, sh in enumerate(opt.get("per_shift") or [], start=6):
                ws_d.cell(i, 1, sh.get("index"))
                ws_d.cell(i, 2, sh.get("shift_type"))
                ws_d.cell(i, 3, sh.get("hours"))
                ws_d.cell(i, 4, sh.get("labour_total"))
                ws_d.cell(i, 5, sh.get("travel_total"))
                ws_d.cell(i, 6, sh.get("meal_total"))
                ws_d.cell(i, 7, sh.get("shift_total"))

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            sheet.column_dimensions[letter].width = max(12, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cost_pdf(
    result: dict[str, Any],
    *,
    title: str = "Cost estimate",
    site: dict[str, Any] | None = None,
    notes: str | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="H1Ven",
            parent=styles["Heading1"],
            textColor=colors.HexColor("#0B3D2E"),
            fontSize=16,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Best",
            parent=styles["Normal"],
            textColor=colors.HexColor("#0B3D2E"),
            fontSize=11,
            leading=14,
            spaceAfter=8,
            fontName="Helvetica-Bold",
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodySmall",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            spaceAfter=4,
        )
    )

    story: list[Any] = []
    story.append(Paragraph("WRU TGS Tracker — Traffic cost export", styles["H1Ven"]))
    story.append(Paragraph(title, styles["Heading2"]))
    story.append(
        Paragraph(
            f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["BodySmall"],
        )
    )
    for line in _site_lines(site):
        story.append(Paragraph(line, styles["BodySmall"]))
    if notes:
        story.append(Paragraph(f"Notes: {notes}", styles["BodySmall"]))
    story.append(Spacer(1, 6))

    booking = _booking_rows(result)
    if booking:
        story.append(Paragraph("<b>Booking requirements</b>", styles["BodySmall"]))
        for b in booking:
            story.append(Paragraph(f"• {b}", styles["BodySmall"]))
        story.append(Spacer(1, 6))

    mode = result.get("mode")
    if mode == "standard":
        echo = result.get("inputs_echo") or {}
        per = result.get("per_shift") or {}
        allow = per.get("allowances") or {}
        data = [
            ["Item", "Amount"],
            ["Shift", f"{echo.get('shift_type')} · {echo.get('shift_hours')}h × {echo.get('total_shifts')} shifts"],
            ["Per-shift pack labour", _money(per.get("shift_labour_total"))],
            ["Per-shift travel", _money(allow.get("travel_total"))],
            ["Per-shift meals", _money(allow.get("meal_total"))],
            ["Per-shift total", _money(per.get("shift_total"))],
            ["Site crew total", _money(result.get("site_crew_total"))],
            ["VMS total", _money((result.get("vms") or {}).get("vms_total"))],
            ["Site traffic total", _money(result.get("site_traffic_total"))],
        ]
        table = Table(data, colWidths=[90 * mm, 70 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D2E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]
            )
        )
        story.append(table)

        lines = per.get("lines") or []
        if lines:
            story.append(Spacer(1, 10))
            story.append(Paragraph("<b>Best pack mix (per shift)</b>", styles["BodySmall"]))
            pdata = [["Qty", "Booking", "Line total"]]
            for line in lines:
                pdata.append(
                    [
                        str(line.get("quantity")),
                        str(line.get("booking_label") or line.get("name")),
                        _money(line.get("line_total")),
                    ]
                )
            pt = Table(pdata, colWidths=[20 * mm, 100 * mm, 40 * mm])
            pt.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E4A")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.append(pt)

    elif mode == "closure_24h":
        rec = result.get("recommendation") or {}
        story.append(
            Paragraph(
                f"BEST OPTION: {rec.get('best_label') or rec.get('cheaper')}",
                styles["Best"],
            )
        )
        story.append(Paragraph(rec.get("summary") or "", styles["BodySmall"]))
        story.append(Spacer(1, 8))

        data = [
            [
                "Pattern",
                "Shifts",
                "Pack labour",
                "Travel",
                "Meals",
                "VMS",
                "Grand total",
            ]
        ]
        best_row = None
        for idx, key in enumerate(("option_3x8", "option_2x12"), start=1):
            opt = result.get(key) or {}
            label = "8-hour shifts" if key == "option_3x8" else "12-hour shifts"
            if opt.get("is_best"):
                best_row = idx
                label = f"★ {label} (BEST)"
            data.append(
                [
                    label,
                    str(opt.get("shifts_required")),
                    _money(opt.get("pack_labour_total")),
                    _money(opt.get("travel_total")),
                    _money(opt.get("meal_total")),
                    _money(opt.get("vms_total")),
                    _money(opt.get("grand_total")),
                ]
            )
        table = Table(data, colWidths=[42 * mm, 16 * mm, 28 * mm, 24 * mm, 24 * mm, 24 * mm, 28 * mm])
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D2E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
        if best_row:
            style_cmds.append(
                ("BACKGROUND", (0, best_row), (-1, best_row), colors.HexColor("#C8E6C9"))
            )
            style_cmds.append(("FONTNAME", (0, best_row), (-1, best_row), "Helvetica-Bold"))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                "Travel applies to TCs, TMA drivers and spotters every shift. "
                "Meals apply only when shift length exceeds the meal threshold (default 9.5h) — "
                "so 12h shifts include meals while 8h shifts usually do not.",
                styles["BodySmall"],
            )
        )

    doc.build(story)
    return buf.getvalue()
