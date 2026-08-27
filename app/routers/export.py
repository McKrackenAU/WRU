from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from ..calculations import COUNCIL_NO_OBJECTION_BUSINESS_DAYS
from ..database import get_db
from ..models import Site
from ..services import site_to_dict
from ..stage_registry import stage_labels_map

router = APIRouter(prefix="/api/export", tags=["export"])

SORT_KEYS = {"pri", "road", "site", "program", "start", "council", "moa"}


@dataclass(frozen=True)
class ClientListView:
    limit: int | None = None
    sort: str | None = None
    direction: str = "asc"
    priorities: set[str] | None = None
    programs: set[str] | None = None


def client_list_view(
    limit: int | None = Query(default=None, ge=1, le=500),
    sort: str | None = Query(default=None),
    direction: str = Query(default="asc", alias="dir"),
    priority: list[str] | None = Query(default=None),
    program: list[str] | None = Query(default=None),
) -> ClientListView:
    key = (sort or "").strip().lower() or None
    if key not in SORT_KEYS:
        key = None
    return ClientListView(
        limit=limit,
        sort=key,
        direction="desc" if str(direction).lower() == "desc" else "asc",
        priorities=None if priority is None else {p for p in priority if p != ""},
        programs=None if program is None else {p for p in program if p != ""},
    )


def _program_key(row: dict) -> str:
    return (row.get("program") or "").strip() or "Unassigned"


def _sort_value(row: dict, key: str):
    if key == "pri":
        try:
            return int(row.get("priority") or 99)
        except (TypeError, ValueError):
            return 99
    if key == "road":
        return (row.get("road_name") or "").lower()
    if key == "site":
        return (row.get("site_number") or "").lower()
    if key == "program":
        return _program_key(row).lower()
    if key == "start":
        return row.get("indicative_start") or "9999-99-99"
    if key == "council":
        wait = row.get("business_days_waiting")
        if wait == "" or wait is None:
            return float("-inf")
        try:
            return float(wait)
        except (TypeError, ValueError):
            return float("-inf")
    if key == "moa":
        return (row.get("moa_number") or "").lower()
    return ""


def apply_client_list_view(rows: list[dict], view: ClientListView | None = None) -> list[dict]:
    """Filter / sort / cap a client list the same way the lists page does."""
    view = view or ClientListView()
    out = list(rows)
    if view.priorities is not None:
        out = [r for r in out if str(r.get("priority") or "") in view.priorities]
    if view.programs is not None:
        out = [r for r in out if _program_key(r) in view.programs]
    if view.sort:
        reverse = view.direction == "desc"

        def decorated(row: dict):
            val = _sort_value(row, view.sort)
            numeric = isinstance(val, (int, float)) and not isinstance(val, bool)
            road = (row.get("road_name") or "").lower()
            return (0 if numeric else 1, val, road)

        out.sort(key=decorated, reverse=reverse)
    else:
        out.sort(key=lambda r: (r.get("priority") or 99, r.get("rank") if r.get("rank") is not None else 999999))
    if view.limit is not None:
        out = out[: view.limit]
    return out


def _view_caption(view: ClientListView | None, team_label: str, count: int) -> str:
    if view and view.limit is not None:
        return f"{team_label} · top {view.limit} · {count} site(s)"
    return f"{team_label} · {count} site(s)"


def _client_list_rows(db: Session, *, team: str, view: ClientListView | None = None) -> list[dict]:
    """team: permits | trims"""
    labels = stage_labels_map(db)
    sites = db.query(Site).filter(Site.archived.is_(False)).all()
    rows = []
    for site in sites:
        data = site_to_dict(site, db=db)
        metrics = data["metrics"]
        flag = (
            metrics.get("on_permits_priority_list")
            if team == "permits"
            else metrics.get("on_trims_priority_list")
        )
        if not flag:
            continue
        current = metrics.get("current_stage")
        council_bits = []
        for c in metrics.get("councils") or []:
            wait = c.get("business_days_waiting")
            council_bits.append(
                f"{c['council_name']}: {c.get('status_label')}"
                + (f" [{wait} bus. days]" if wait is not None and c.get("status") == "waiting" else "")
            )
        rows.append(
            {
                "team": team,
                "priority": data["today_priority"],
                "rank": metrics.get("permits_priority_rank"),
                "road_name": site.road_name,
                "site_number": site.site_number,
                "program": site.program or "",
                "councils": "; ".join(data["councils"]),
                "council_status": " | ".join(council_bits),
                "business_days_waiting": metrics.get("max_council_business_days_waiting")
                if metrics.get("max_council_business_days_waiting") is not None
                else "",
                "indicative_start": site.indicative_site_start_date.isoformat()
                if site.indicative_site_start_date
                else "",
                "days_to_start": metrics.get("days_to_start")
                if metrics.get("days_to_start") is not None
                else "",
                "moa_must_have": site.moa_must_have_received_date.isoformat()
                if site.moa_must_have_received_date
                else "",
                "must_have_status": metrics.get("must_have_status", {}).get("label", ""),
                "moa_number": site.moa_number or "",
                "moa_submission_date": site.moa_submission_date.isoformat()
                if site.moa_submission_date
                else "",
                "tgs_reference": site.tgs_reference or "",
                "current_stage": labels.get(current or "", current or ""),
                "progress_pct": metrics.get("workflow_progress_pct", 0),
                "client_list": metrics.get("client_list"),
                "is_generic_moa": "yes" if data.get("is_generic_moa") else "",
                "comments": (site.comments or "").replace("\n", " "),
            }
        )
    return apply_client_list_view(rows, view)


# Permits / TRIMS client-list export: six columns, matching the printed list.
CLIENT_COLUMNS = [
    ("priority", "Priority"),
    ("moa_number", "MoA Number"),
    ("road_name", "Road"),
    ("site_number", "Site"),
    ("program", "Program"),
    ("indicative_start", "Indicative start"),
]
CLIENT_FIELDS = [key for key, _ in CLIENT_COLUMNS]
CLIENT_HEADERS = [label for _, label in CLIENT_COLUMNS]


def _client_export_values(row: dict) -> dict[str, Any]:
    return {label: row.get(key, "") for key, label in CLIENT_COLUMNS}


def _csv_response(rows: list[dict], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CLIENT_HEADERS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(_client_export_values(row))
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_bytes(rows: list[dict], sheet_name: str, title: str, view: ClientListView | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.cell(1, 1, title).font = Font(bold=True, size=13)
    cap = "Exported " + date.today().isoformat()
    if view and view.limit is not None:
        cap += f" · top {view.limit}"
    cap += (
        f" · Council no-objection assumed after "
        f"{COUNCIL_NO_OBJECTION_BUSINESS_DAYS} business days without response"
    )
    ws.cell(2, 1, cap)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3D2E")
    for col, h in enumerate(CLIENT_HEADERS, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, row in enumerate(rows, start=5):
        values = _client_export_values(row)
        for c_idx, header in enumerate(CLIENT_HEADERS, 1):
            ws.cell(r_idx, c_idx, values.get(header, ""))
    for col in ws.columns:
        letter = col[0].column_letter
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 48)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_cell(text: Any, style: ParagraphStyle) -> Paragraph:
    raw = "" if text is None else str(text)
    safe = (
        raw.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(safe or "—", style)


def _pdf_bytes(rows: list[dict], *, title: str, team_label: str, view: ClientListView | None = None) -> bytes:
    """Landscape A4 priority list with Ventia / VenInspect-style branding."""
    from ..pdf_brand import (
        GREEN,
        MUTED,
        ROW_ALT,
        RULE,
        WARN_TINT,
        branded_margins,
        draw_branded_page,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title=title,
        **branded_margins(landscape_mode=True),
    )
    caption = _view_caption(view, team_label, len(rows))
    doc.brand_eyebrow = "PRIORITY LIST"
    doc.brand_title = title
    doc.brand_subtitle = (
        f"{caption} · "
        f"Council no-objection after {COUNCIL_NO_OBJECTION_BUSINESS_DAYS} business days"
    )
    doc.brand_doc_kind = "Priority List"
    doc.brand_product = "WRU TGS TRACKER"
    doc.brand_footer_meta = caption

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Th",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.5,
            leading=9,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Td",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#1a1a1a"),
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="TdCenter",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#1a1a1a"),
            alignment=TA_CENTER,
        )
    )
    styles.add(
        ParagraphStyle(
            name="FootNote",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=MUTED,
            spaceBefore=6,
        )
    )

    headers = [
        "Pri",
        "MoA #",
        "Road",
        "Site",
        "Program",
        "Indicative start",
    ]
    data: list[list[Any]] = [[_pdf_cell(h, styles["Th"]) for h in headers]]

    if not rows:
        data.append(
            [_pdf_cell("No sites on this priority list.", styles["Td"])]
            + [""] * (len(headers) - 1)
        )
    else:
        for row in rows:
            data.append(
                [
                    _pdf_cell(row.get("priority"), styles["TdCenter"]),
                    _pdf_cell(row.get("moa_number"), styles["Td"]),
                    _pdf_cell(row.get("road_name"), styles["Td"]),
                    _pdf_cell(row.get("site_number"), styles["Td"]),
                    _pdf_cell(row.get("program"), styles["Td"]),
                    _pdf_cell(row.get("indicative_start"), styles["Td"]),
                ]
            )

    # Landscape A4 with 12mm margins ≈ 273mm usable
    col_widths = [
        16 * mm,  # Pri
        42 * mm,  # MoA #
        78 * mm,  # Road
        32 * mm,  # Site
        64 * mm,  # Program
        41 * mm,  # Indicative start
    ]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), GREEN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, RULE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
    ]
    for i, row in enumerate(rows, start=1):
        if str(row.get("priority")) == "1":
            style_cmds.append(("BACKGROUND", (0, i), (0, i), WARN_TINT))
    if not rows:
        style_cmds.append(("SPAN", (0, 1), (-1, 1)))
    table.setStyle(TableStyle(style_cmds))

    story: list[Any] = [
        table,
        Spacer(1, 4 * mm),
        Paragraph(
            "Ventia — confidential priority list for client coordination.",
            styles["FootNote"],
        ),
    ]
    doc.build(story, onFirstPage=draw_branded_page, onLaterPages=draw_branded_page)
    return buf.getvalue()


@router.get("/priority-list.csv")
def export_permits_priority_csv(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    """Legacy alias — Permits team client list."""
    rows = _client_list_rows(db, team="permits", view=view)
    return _csv_response(rows, f"WRU_Permits_priority_{date.today().isoformat()}.csv")


@router.get("/permits-list.csv")
def export_permits_csv(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="permits", view=view)
    return _csv_response(rows, f"WRU_Permits_list_{date.today().isoformat()}.csv")


@router.get("/trims-list.csv")
def export_trims_csv(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="trims", view=view)
    return _csv_response(rows, f"WRU_TRIMS_list_{date.today().isoformat()}.csv")


@router.get("/permits-list.xlsx")
def export_permits_xlsx(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="permits", view=view)
    data = _xlsx_bytes(rows, "Permits", "WRU TGS Tracker — Permits team priority list", view)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="WRU_Permits_list_{date.today().isoformat()}.xlsx"'
        },
    )


@router.get("/trims-list.xlsx")
def export_trims_xlsx(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="trims", view=view)
    data = _xlsx_bytes(rows, "TRIMS", "WRU TGS Tracker — TRIMS team priority list", view)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="WRU_TRIMS_list_{date.today().isoformat()}.xlsx"'
        },
    )


@router.get("/permits-list.pdf")
def export_permits_pdf(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="permits", view=view)
    data = _pdf_bytes(
        rows,
        title="DTP — Permits priority list",
        team_label="Permits team",
        view=view,
    )
    return Response(
        content=data,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="WRU_Permits_list_{date.today().isoformat()}.pdf"'
        },
    )


@router.get("/trims-list.pdf")
def export_trims_pdf(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    rows = _client_list_rows(db, team="trims", view=view)
    data = _pdf_bytes(
        rows,
        title="DTP — TRIMS priority list",
        team_label="TRIMS team",
        view=view,
    )
    return Response(
        content=data,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="WRU_TRIMS_list_{date.today().isoformat()}.pdf"'
        },
    )


@router.get("/priority-list.pdf")
def export_permits_priority_pdf(
    view: ClientListView = Depends(client_list_view),
    db: Session = Depends(get_db),
):
    """Legacy alias — Permits team PDF."""
    return export_permits_pdf(view=view, db=db)


@router.get("/sites.csv")
def export_sites_csv(archived: bool = False, db: Session = Depends(get_db)):
    labels = stage_labels_map(db)
    sites = db.query(Site).filter(Site.archived.is_(archived)).all()
    buf = io.StringIO()
    fieldnames = [
        "id",
        "road_name",
        "site_number",
        "program",
        "councils",
        "council_status",
        "business_days_waiting",
        "financial_year",
        "priority",
        "indicative_start",
        "moa_must_have",
        "moa_number",
        "current_stage",
        "progress_pct",
        "client_list",
        "on_permits_priority_list",
        "on_trims_priority_list",
        "is_generic_moa",
        "archived",
        "comments",
    ]
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for site in sites:
        data = site_to_dict(site, db=db)
        m = data["metrics"]
        council_bits = [f"{c['council_name']}: {c.get('status_label')}" for c in (m.get("councils") or [])]
        writer.writerow(
            {
                "id": site.id,
                "road_name": site.road_name,
                "site_number": site.site_number,
                "program": site.program or "",
                "councils": "; ".join(data["councils"]),
                "council_status": " | ".join(council_bits),
                "business_days_waiting": m.get("max_council_business_days_waiting")
                if m.get("max_council_business_days_waiting") is not None
                else "",
                "financial_year": data["financial_year"],
                "priority": data["today_priority"],
                "indicative_start": site.indicative_site_start_date or "",
                "moa_must_have": site.moa_must_have_received_date or "",
                "moa_number": site.moa_number or "",
                "current_stage": labels.get(m.get("current_stage") or "", m.get("current_stage") or ""),
                "progress_pct": m.get("workflow_progress_pct", 0),
                "client_list": m.get("client_list"),
                "on_permits_priority_list": m.get("on_permits_priority_list"),
                "on_trims_priority_list": m.get("on_trims_priority_list"),
                "is_generic_moa": data.get("is_generic_moa"),
                "archived": site.archived,
                "comments": (site.comments or "").replace("\n", " "),
            }
        )
    kind = "archive" if archived else "active"
    filename = f"WRU_TGS_{kind}_sites_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
