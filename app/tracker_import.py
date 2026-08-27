"""Import sites from WRU Traffic TGS-MOA Tracker Excel (.xlsx / .xlsm)."""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .calculations import expand_workflow_prefix
from .lookups import sync_usage_into_lookups
from .models import Site
from .services import apply_workflow, ensure_workflow_steps, set_councils, sync_computed_fields
from .stage_registry import active_programs, stage_keys as registry_stage_keys

# Spreadsheet status text → cumulative completed stages
STATUS_STAGES: dict[str, list[str]] = {
    "not yet started": [],
    "tgs markup complete": ["tgs_markup_completed"],
    "submitted to tmd": ["tgs_markup_completed", "submitted_to_tmd"],
    "plan received": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
    ],
    "ready to submit moa": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
    ],
    "moa submitted": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
        "moa_submitted",
    ],
    "moa with trims": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
        "moa_submitted",
        "moa_with_trims",
    ],
    "revision needed": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
        "moa_submitted",
        "moa_with_trims",
        "revision_needed",
    ],
    "moa received": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
        "moa_submitted",
        "moa_with_trims",
        "moa_received",
        "ready_for_works",
    ],
    "ready for works": [
        "tgs_markup_completed",
        "submitted_to_tmd",
        "ventia_review",
        "plan_received",
        "ready_to_submit_moa",
        "moa_submitted",
        "moa_with_trims",
        "moa_received",
        "ready_for_works",
    ],
}

# Spreadsheet dropdown / formula casing variants → canonical STATUS_STAGES key
STATUS_ALIASES: dict[str, str] = {
    "not yet started": "not yet started",
    "not started": "not yet started",
    "tgs markup completed": "tgs markup complete",
    "tgs markup complete": "tgs markup complete",
    "markup complete": "tgs markup complete",
    "submitted to tmd": "submitted to tmd",
    "submitted to tm": "submitted to tmd",
    "submitted to traffic management": "submitted to tmd",
    "submitted to traffic management designer": "submitted to tmd",
    "plan received": "plan received",
    "ready to submit moa": "ready to submit moa",
    "ready for submit moa": "ready to submit moa",
    "moa submitted": "moa submitted",
    "moa with trims": "moa with trims",
    'moa "with trims"': "moa with trims",
    "moa with trims - (remove from permits priority list)": "moa with trims",
    "revision needed": "revision needed",
    "revision required": "revision needed",
    "moa received": "moa received",
    "moa received / approved": "moa received",
    "moa approved": "moa received",
    "ready for works": "ready for works",
    "ready to works": "ready for works",
    "ventia review": "plan received",
}

# Back-compat name used by tests / callers
STATUS_MAP: list[tuple[str, list[str]]] = [(k, v) for k, v in STATUS_STAGES.items()]

WORKFLOW_KEYS = [
    "tgs_markup_completed",
    "submitted_to_tmd",
    "plan_received",
    "ready_to_submit_moa",
    "moa_submitted",
    "moa_with_trims",
    "revision_needed",
    "moa_received",
    "ready_for_works",
]

# Columns H–O after the master status dropdown (G)
LAMP_KEYS = WORKFLOW_KEYS[1:]

STAGE_TO_STATUS = {
    "tgs_markup_completed": "tgs markup complete",
    "submitted_to_tmd": "submitted to tmd",
    "plan_received": "plan received",
    "ready_to_submit_moa": "ready to submit moa",
    "moa_submitted": "moa submitted",
    "moa_with_trims": "moa with trims",
    "revision_needed": "revision needed",
    "moa_received": "moa received",
    "ready_for_works": "ready for works",
}

# Column A section headers in the V6 workbook
SECTION_HINTS = [
    ("lcp - fmrp", "LCP-FMRP"),
    ("lcp-fmrp", "LCP-FMRP"),
    ("fmrp non", "FMRP Non-Commit"),
    ("maintenace", "LCP Maintenance Misc"),  # typo in sheet
    ("maintenance", "LCP Maintenance Misc"),
    ("structure", "Structures"),
    ("generic", "Generics MTMP/ITMP"),
    ("routine", "Routine Maintenance"),
]

SKIP_ROAD = {
    "road name",
    "add new line above",
    "totals",
    "jobs",
    "none",
    "completed",
    "not completed",
    "% complete",
    "",
}

SKIP_SITE = {
    "completed",
    "not completed",
    "% complete",
    "%complete",
}

# V6 column letters (1-based) when headers cannot be read
DEFAULT_COLS = {
    "include": 1,
    "road": 2,
    "site": 3,
    "start": 4,
    "must_have": 5,
    "status": 7,
    "comments": 16,
    "moa_number": 18,
    "moa_sub": 19,
    "moa_rec": 20,
    "moa_start": 22,
    "moa_exp": 23,
    "council": 24,
    "council_sub": 25,
    "council_obj": 26,
    "ext_flag": 28,
    "ext_sub": 29,
    "ext_rec": 30,
    "ext_start": 32,
    "ext_exp": 33,
    "job_date": 34,
    "job_done": 35,
}

COUNCIL_ALIASES = {
    "brim": "Brimbank",
    "brimbank": "Brimbank",
    "mel": "Melbourne",
    "melbourne": "Melbourne",
    "hob": "Hobsons Bay",
    "hobs": "Hobsons Bay",
    "hobsons": "Hobsons Bay",
    "hobsons bay": "Hobsons Bay",
    "hobsons bay city": "Hobsons Bay",
    "marib": "Maribyrnong",
    "maribyrnong": "Maribyrnong",
    "maribynong": "Maribyrnong",
    "melton": "Melton",
    "wyd": "Wyndham",
    "wyndham": "Wyndham",
}

TGS_REF_RE = re.compile(r"\b(TGS[-A-Z0-9]{3,}(?:-[A-Z0-9]+)*)", re.I)

DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d/%m/%y",
    "%d-%m-%y",
    "%d-%b-%y",
    "%d-%b-%Y",
    "%d-%B-%y",
    "%d-%B-%Y",
    "%d/%b/%y",
    "%d/%b/%Y",
    "%m/%d/%Y",
)


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _comment_text(value: Any) -> str | None:
    """Keep line breaks in the Comments column (ops correspondence log)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"[ \t]+", " ", line).rstrip() for line in text.split("\n")]
    text = "\n".join(lines).strip()
    return text or None


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        # Excel sometimes stores small serials as datetime near 1900
        if value.year < 1950:
            return None
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm(value)
    if not text or text.lower() in ("received", "n/a", "na", "-", "yes", "no", "not yet submitted"):
        return None
    token = text.split()[0]
    for candidate in (text, token, text[:10]):
        for fmt in DATE_FORMATS:
            try:
                parsed = datetime.strptime(candidate, fmt)
                if parsed.year < 1950:
                    continue
                return parsed.date()
            except ValueError:
                continue
    try:
        n = float(value)
        if 20000 < n < 80000:
            return (datetime(1899, 12, 30) + timedelta(days=int(n))).date()
    except (TypeError, ValueError):
        pass
    return None


def _canonical_council(name: str) -> str:
    key = re.sub(r"\s+", " ", name or "").strip().lower()
    if not key:
        return ""
    return COUNCIL_ALIASES.get(key, re.sub(r"\s+", " ", name).strip())


def _parse_council_blob(text: str) -> list[dict[str, Any]]:
    """Parse 'Brim-30/07/25' / multi-line no-objection notes into council rows."""
    rows: list[dict[str, Any]] = []
    raw = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    for line in re.split(r"[\n;]+", raw):
        line = line.strip()
        if not line or line in "-–":
            continue
        m = re.match(
            r"^([A-Za-z][A-Za-z .']*?)\s*[-–:]\s*(\d{1,2}[/-][A-Za-z0-9]{1,9}[/-]\d{2,4})?\s*$",
            line,
        )
        if m:
            name = _canonical_council(m.group(1))
            rows.append({"council_name": name, "no_objection_date": _as_date(m.group(2))})
            continue
        dt = _as_date(line)
        if dt:
            rows.append({"council_name": "", "no_objection_date": dt})
    return rows


def _councils_from_cells(x_val: Any, y_val: Any, z_val: Any) -> list[dict[str, Any]]:
    names = [
        _canonical_council(part)
        for part in re.split(r"[,;/]", _norm(x_val))
        if part.strip()
    ]
    names = [
        n
        for n in names
        if n and n.lower() not in ("multiple", "multi", "various", "n/a", "na", "none")
    ]
    submitted = _as_date(y_val)
    objection = _as_date(z_val)
    blob_rows: list[dict[str, Any]] = []
    if objection is None and z_val not in (None, ""):
        blob_rows = [r for r in _parse_council_blob(str(z_val)) if r.get("council_name")]
    if blob_rows:
        return [
            {
                "council_name": r["council_name"],
                "submitted_to_council_date": submitted,
                "no_objection_date": r.get("no_objection_date"),
            }
            for r in blob_rows
        ]
    return [
        {
            "council_name": name,
            "submitted_to_council_date": submitted,
            "no_objection_date": objection,
        }
        for name in names
    ]


def _tgs_reference(*parts: Any) -> str | None:
    for part in parts:
        text = part if isinstance(part, str) else _comment_text(part) or _norm(part)
        if not text:
            continue
        match = TGS_REF_RE.search(text)
        if match:
            return match.group(1).upper() if match.group(1).upper().startswith("TGS") else match.group(1)
    return None


def _is_junk_row(road: str, site: str, status: Any) -> bool:
    road_l = road.lower()
    site_l = site.lower()
    if road_l in SKIP_ROAD or site_l in SKIP_SITE:
        return True
    if road_l.startswith("totals") or road_l.startswith("add new"):
        return True
    if re.fullmatch(r"\d+(\.\d+)?", road):
        return True
    if "%" in road_l or "%" in site_l:
        return True
    if isinstance(status, (int, float)) and not isinstance(status, bool):
        return True
    if isinstance(status, datetime) and status.year < 1950:
        return True
    return False


def _cell(ws, row: int, cols: dict[str, int], key: str):
    idx = cols.get(key)
    if not idx:
        return None
    return ws.cell(row, idx).value


def _header_label(value: Any) -> str:
    return re.sub(r"\s+", " ", _norm(value).lower().replace("\n", " "))


def _columns_from_header(ws) -> tuple[int, dict[str, int]]:
    """Return (header_row, column map). Falls back to V6 letters."""
    mapping = dict(DEFAULT_COLS)
    header_row = 1
    for r in range(1, 8):
        labels = [_header_label(ws.cell(r, c).value) for c in range(1, 46)]
        if "road name" in labels:
            header_row = r
            seen_start = 0
            seen_exp = 0
            seen_wait = 0
            seen_job = 0
            for c in range(1, 46):
                h = _header_label(ws.cell(r, c).value)
                if not h:
                    continue
                if h == "road name":
                    mapping["road"] = c
                elif h == "site number":
                    mapping["site"] = c
                elif "indicative" in h and "start" in h:
                    mapping["start"] = c
                elif "must have" in h:
                    mapping["must_have"] = c
                elif "tgs" in h and "markup" in h:
                    mapping["status"] = c
                elif h == "comments":
                    mapping["comments"] = c
                elif h == "moa number":
                    mapping["moa_number"] = c
                elif "moa submission" in h:
                    mapping["moa_sub"] = c
                elif h.startswith("moa received date"):
                    mapping["moa_rec"] = c
                elif "council submission" in h:
                    mapping["council_sub"] = c
                elif "no objection recieved" in h or "no objection received" in h:
                    mapping["council_obj"] = c
                elif h == "council":
                    mapping["council"] = c
                elif "extension or change" in h or "change requried" in h:
                    mapping["ext_flag"] = c
                elif "extension/change submission" in h or "e/c submission" in h:
                    mapping["ext_sub"] = c
                elif "e/c received" in h:
                    mapping["ext_rec"] = c
                elif "moa start" in h:
                    seen_start += 1
                    mapping["moa_start" if seen_start == 1 else "ext_start"] = c
                elif "moa expiry" in h:
                    seen_exp += 1
                    mapping["moa_exp" if seen_exp == 1 else "ext_exp"] = c
                elif "job completed" in h and "date" in h:
                    mapping["job_date"] = c
                    seen_job += 1
                elif h == "job completed" or (h.startswith("job completed") and "date" not in h):
                    mapping["job_done"] = c
                elif "include" in h:
                    mapping["include"] = c
            break
    return header_row, mapping


def _row_excel_comment(ws, row: int, cols: dict[str, int]) -> str | None:
    bits: list[str] = []
    for key in ("comments", "moa_rec", "moa_number", "status"):
        idx = cols.get(key)
        if not idx:
            continue
        cmt = ws.cell(row, idx).comment
        if cmt and cmt.text:
            bits.append(_comment_text(cmt.text) or "")
    bits = [b for b in bits if b]
    return "\n".join(bits) if bits else None


def _canonical_status_key(status: str) -> str | None:
    """Map spreadsheet status text to a STATUS_STAGES key.

    Exact match first. Then the longest alias contained in the cell (so
    'MoA Submitted - TRIMS requested…' still maps). Never match a short
    cell inside a longer alias ('received' must not become 'plan received').
    """
    key = _norm(status).lower()
    if not key:
        return None
    if key in STATUS_ALIASES:
        return STATUS_ALIASES[key]
    contained = [alias for alias in STATUS_ALIASES if alias in key]
    if not contained:
        return None
    contained.sort(key=len, reverse=True)
    return STATUS_ALIASES[contained[0]]


def _lamp_reached(val: Any) -> bool | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime) and val.year < 1950:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val) != 0
    n = _norm(val).lower()
    if n in ("0", "no", "n", "false", "-"):
        return False
    if n in ("yes", "y", "1", "true"):
        return True
    return None


def _furthest_from_lamps(values: list[Any] | None) -> str | None:
    """V6 H–O lamps: blank = passed, 0 = not yet reached, Yes = passed (legacy)."""
    if not values:
        return None
    padded = list(values[: len(LAMP_KEYS)]) + [None] * max(0, len(LAMP_KEYS) - len(values))
    if all(v in (None, "") for v in padded):
        return None
    furthest: str | None = None
    for key, val in zip(LAMP_KEYS, padded):
        if val in (None, ""):
            furthest = key
            continue
        reached = _lamp_reached(val)
        if reached is False:
            break
        if reached is True:
            furthest = key
    return furthest


def _flags_for_status_key(canonical: str) -> dict[str, bool]:
    stages = set(STATUS_STAGES.get(canonical, []))
    flags = {k: k in stages for k in WORKFLOW_KEYS}
    flags["ventia_review"] = "ventia_review" in stages
    if canonical == "revision needed":
        flags["revision_needed"] = True
    return flags


def _status_workflow(status: str, lamps: list[Any] | None = None) -> tuple[dict[str, bool], str | None]:
    """Return (workflow flags, unmatched original text if we fell back)."""
    key = _norm(status).lower()
    unmatched: str | None = None
    canonical: str | None = None

    if not key or key in ("n/a", "na", "-", "0"):
        canonical = None
    elif key in ("yes", "y"):
        lamp_stage = _furthest_from_lamps(lamps)
        if lamp_stage:
            canonical = STAGE_TO_STATUS.get(lamp_stage)
        else:
            canonical = "ready for works"
    elif key in ("no", "n"):
        canonical = "not yet started"
    elif key.isdigit():
        unmatched = status
        canonical = None
    else:
        canonical = _canonical_status_key(status)
        if canonical is None:
            unmatched = status

    if canonical is None:
        lamp_stage = _furthest_from_lamps(lamps)
        if lamp_stage:
            canonical = STAGE_TO_STATUS.get(lamp_stage)
            unmatched = None
        else:
            canonical = "not yet started"

    return _flags_for_status_key(canonical or "not yet started"), unmatched


def _section_from_a(value: Any) -> str | None:
    text = _norm(value).lower()
    if not text:
        return None
    for hint, name in SECTION_HINTS:
        if hint in text:
            return name
    return None


def _find_tracker_sheet(wb):
    for name in wb.sheetnames:
        if "tgs" in name.lower() and "moa" in name.lower():
            return wb[name]
    for name in wb.sheetnames:
        if "enable" in name.lower() or "dashboard" in name.lower():
            continue
        return wb[name]
    return wb[wb.sheetnames[0]]


def parse_tracker_workbook(content: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    ws = _find_tracker_sheet(wb)
    header_row, cols = _columns_from_header(ws)
    rows_out: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    unmatched_statuses: list[str] = []
    current_program = "LCP-FMRP"

    for r in range(1, (ws.max_row or 0) + 1):
        if r == header_row:
            continue
        a_raw = _cell(ws, r, cols, "include")
        section = _section_from_a(a_raw)
        if section:
            current_program = section
            continue

        b = _norm(_cell(ws, r, cols, "road"))
        c = _norm(_cell(ws, r, cols, "site"))
        status_raw = _cell(ws, r, cols, "status")
        if not b:
            continue
        if _is_junk_row(b, c, status_raw):
            skipped.append({"row": r, "road_name": b, "site_number": c, "reason": "totals/placeholder"})
            continue

        status = _norm(status_raw)
        status_idx = cols.get("status") or 7
        lamps = [ws.cell(r, status_idx + i).value for i in range(1, 9)]
        workflow, unmatched = _status_workflow(status, lamps)
        if unmatched:
            unmatched_statuses.append(_norm(unmatched))

        comments = _comment_text(_cell(ws, r, cols, "comments"))
        extra_cmt = _row_excel_comment(ws, r, cols)
        if extra_cmt:
            comments = f"{comments}\n{extra_cmt}".strip() if comments else extra_cmt

        moa_raw = _cell(ws, r, cols, "moa_number")
        moa_number = _comment_text(moa_raw)
        if moa_number:
            moa_number = " / ".join(
                part.strip() for part in moa_number.split("\n") if part.strip()
            )
        tgs = _tgs_reference(comments, moa_number, b, c)

        site_number = c or (tgs or "") or re.sub(r"[^A-Za-z0-9]+", "-", b)[:48].strip("-") or f"ROW-{r}"
        site_number = site_number[:64]

        councils = _councils_from_cells(
            _cell(ws, r, cols, "council"),
            _cell(ws, r, cols, "council_sub"),
            _cell(ws, r, cols, "council_obj"),
        )
        ext = _norm(_cell(ws, r, cols, "ext_flag")) or "No"
        if ext.lower() not in ("yes", "no", "n/a", "na"):
            ext = "No"
        if ext.lower() == "na":
            ext = "N/A"
        job_done = _norm(_cell(ws, r, cols, "job_done")).lower()
        include_in_totals = False
        try:
            include_in_totals = int(a_raw) == 1
        except (TypeError, ValueError):
            include_in_totals = _norm(a_raw).lower() in ("1", "yes", "y", "true")

        start = _as_date(_cell(ws, r, cols, "start"))
        rows_out.append(
            {
                "road_name": b,
                "site_number": site_number,
                "program": current_program,
                "tgs_reference": tgs,
                "indicative_site_start_date": start,
                "moa_must_have_received_date": None,
                "must_have_manual": False,
                "comments": comments,
                "moa_number": moa_number,
                "moa_submission_date": _as_date(_cell(ws, r, cols, "moa_sub")),
                "moa_received_date": _as_date(_cell(ws, r, cols, "moa_rec")),
                "moa_start_date": _as_date(_cell(ws, r, cols, "moa_start")),
                "moa_expiry_date": _as_date(_cell(ws, r, cols, "moa_exp")),
                "councils": councils,
                "extension_flag": ext.title() if ext.lower() != "n/a" else "N/A",
                "extension_submission_date": _as_date(_cell(ws, r, cols, "ext_sub")),
                "extension_received_date": _as_date(_cell(ws, r, cols, "ext_rec")),
                "extension_start_date": _as_date(_cell(ws, r, cols, "ext_start")),
                "extension_expiry_date": _as_date(_cell(ws, r, cols, "ext_exp")),
                "job_completed_date": _as_date(_cell(ws, r, cols, "job_date")),
                "include_in_totals": include_in_totals,
                "is_generic_moa": "generic" in (current_program or "").lower(),
                "workflow": workflow,
                "status_text": status,
                "status_unmatched": unmatched,
                "archive": job_done in ("yes", "y", "true", "1"),
                "sheet_row": r,
            }
        )
    sheet_name = ws.title
    wb.close()
    unique_unmatched = sorted({s for s in unmatched_statuses if s})
    return {
        "rows": rows_out,
        "skipped": skipped[:80],
        "unmatched_statuses": unique_unmatched,
        "sheet_name": sheet_name,
        "parsed": len(rows_out),
    }


def import_tracker_rows(
    db: Session,
    rows: list[dict[str, Any]],
    *,
    update_existing: bool = True,
) -> dict[str, Any]:
    programs = active_programs(db)
    created = updated = skipped = archived = 0
    errors: list[str] = []
    order_by_program: dict[str, int] = {}

    for raw in rows:
        try:
            road = raw["road_name"]
            site_no = raw["site_number"]
            existing = (
                db.query(Site)
                .filter(Site.road_name.ilike(road), Site.site_number.ilike(site_no))
                .first()
            )
            program = raw.get("program")
            if program and program not in programs:
                pass

            if existing and not update_existing:
                skipped += 1
                continue

            site = existing or Site(road_name=road, site_number=site_no)
            if not existing:
                db.add(site)
                created += 1
            else:
                updated += 1

            site.program = program or site.program
            site.tgs_reference = raw.get("tgs_reference") or site.tgs_reference
            site.indicative_site_start_date = raw.get("indicative_site_start_date")
            site.must_have_manual = bool(raw.get("must_have_manual"))
            if raw.get("must_have_manual"):
                site.moa_must_have_received_date = raw.get("moa_must_have_received_date")
            site.comments = raw.get("comments")
            site.moa_number = raw.get("moa_number")
            site.moa_submission_date = raw.get("moa_submission_date")
            site.moa_received_date = raw.get("moa_received_date")
            site.moa_start_date = raw.get("moa_start_date")
            site.moa_expiry_date = raw.get("moa_expiry_date")
            site.extension_flag = raw.get("extension_flag")
            site.extension_submission_date = raw.get("extension_submission_date")
            site.extension_received_date = raw.get("extension_received_date")
            site.extension_start_date = raw.get("extension_start_date")
            site.extension_expiry_date = raw.get("extension_expiry_date")
            site.job_completed_date = raw.get("job_completed_date")
            site.include_in_totals = bool(raw.get("include_in_totals", True))
            if raw.get("is_generic_moa"):
                site.is_generic_moa = True
            prog_key = program or site.program or "Unassigned"
            order_by_program[prog_key] = order_by_program.get(prog_key, 0) + 10
            site.register_order = order_by_program[prog_key]

            db.flush()
            ensure_workflow_steps(site, db)
            # Reset leftover steps from earlier imports (e.g. inactive ventia_review).
            full_wf = {step.stage: False for step in site.workflow_steps}
            full_wf.update(raw.get("workflow") or {})
            full_wf = expand_workflow_prefix(full_wf, registry_stage_keys(db))
            apply_workflow(site, full_wf, db)
            set_councils(site, raw.get("councils") or [])
            if raw.get("archive"):
                from datetime import datetime, timezone

                from .services import infer_financial_year

                site.archived = True
                site.archived_at = datetime.now(timezone.utc)
                site.archived_fy = infer_financial_year(site)
                archived += 1
            sync_computed_fields(site, db)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{raw.get('road_name')} / {raw.get('site_number')}: {exc}")

    db.commit()
    sync_usage_into_lookups(db, "road")
    return {
        "parsed": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "archived": archived,
        "errors": errors[:50],
    }
