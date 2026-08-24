"""Parse and import asphalt / traffic rate cards from CSV or Excel."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from .asphalt_engine import apply_stored_rates, infer_rate_type, normalize_unit
from .models import AsphaltRate, AsphaltSubcontractor, LabourRate

ASPHALT_TEMPLATE_HEADERS = [
    "subcontractor",
    "name",
    "unit",
    "rate_type",
    "unit_rate",
    "day_rate",
    "night_rate",
    "weekend_rate",
    "public_holiday_rate",
]

TRAFFIC_TEMPLATE_HEADERS = [
    "name",
    "kind",
    "pack_people",
    "includes_vehicle",
    "day_ordinary",
    "day_overtime",
    "night_ordinary",
    "night_overtime",
    "weekend_ordinary",
    "weekend_overtime",
    "public_holiday_ordinary",
    "public_holiday_overtime",
    "active",
]


def _norm_header(value: Any) -> str:
    return "".join(str(value or "").strip().lower().replace("²", "2").split())


def _as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _as_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("1", "true", "yes", "y", "active")


def _header_map(headers: list[str]) -> dict[str, int]:
    aliases = {
        "subcontractor": ("subcontractor", "subbie", "contractor", "supplier"),
        "name": ("name", "treatment", "item", "description", "mix"),
        "unit": ("unit", "uom"),
        "rate_type": ("ratetype", "type", "kind"),
        "unit_rate": ("unitrate", "rate", "sqmrate", "m2rate"),
        "day_rate": ("dayrate", "day"),
        "night_rate": ("nightrate", "night"),
        "weekend_rate": ("weekendrate", "weekend", "sundayrate", "sunday"),
        "saturday_rate": ("saturdayrate", "saturday", "sat"),
        "public_holiday_rate": ("publicholidayrate", "phrate", "ph", "holiday"),
        "kind": ("kind", "ratekind"),
        "pack_people": ("packpeople", "tcs", "people"),
        "includes_vehicle": ("includesvehicle", "vehicle"),
        "day_ordinary": ("dayordinary", "dayord"),
        "day_overtime": ("dayovertime", "dayot"),
        "night_ordinary": ("nightordinary", "nightord"),
        "night_overtime": ("nightovertime", "nightot"),
        "weekend_ordinary": ("weekendordinary", "weekendord", "sundayordinary", "sunord"),
        "weekend_overtime": ("weekendovertime", "weekendot", "sundayovertime", "sunot"),
        "public_holiday_ordinary": ("publicholidayordinary", "phordinary", "phord"),
        "public_holiday_overtime": ("publicholidayovertime", "phot"),
        "active": ("active",),
    }
    mapped: dict[str, int] = {}
    normalized = [_norm_header(h) for h in headers]
    for key, names in aliases.items():
        for i, h in enumerate(normalized):
            if h in names:
                mapped[key] = i
                break
    return mapped


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if any(str(c).strip() for c in row)]


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if any(v is not None and str(v).strip() for v in values):
            rows.append(values)
    wb.close()
    return rows


def parse_tabular(content: bytes, filename: str) -> list[list[Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _rows_from_csv(content)
    if name.endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx(content)
    raise ValueError("Upload a .csv or .xlsx file")


def build_asphalt_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asphalt rates"
    ws.append(ASPHALT_TEMPLATE_HEADERS)
    examples = [
        ["BORAL", "50mm HP Mill & Resheet", "m2", "unit", 32.71, "", "", "", ""],
        ["RABS", "50mm HP Mill & Resheet", "m2", "unit", 33.89, "", "", "", ""],
        ["PRESTIGE", "50mm HP Mill & Resheet", "m2", "unit", 29.50, "", "", "", ""],
        ["RABS", "Mobilisation", "shift", "shift", "", 2500, 3200, 3800, 4500],
        ["BORAL", "Crew / establishment", "shift", "shift", "", 1800, 2200, 2600, 3100],
    ]
    for row in examples:
        ws.append(row)
    ws.append([])
    ws.append(
        [
            "Tips: shared treatment names (column name) let subcontractors quote the same mix.",
            "Unit items (m2/tonne/lm) only need unit_rate. Shift items (mobilisation/crew) need day/night/weekend/PH.",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_traffic_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Traffic rates"
    ws.append(TRAFFIC_TEMPLATE_HEADERS)
    ws.append(["2 TC + vehicle", "crew_pack", 2, "yes", 130, 170, 160, 210, 190, 250, 220, 280, "yes"])
    ws.append(["TMA", "tma", 0, "yes", 180, 220, 210, 260, 240, 300, 280, 340, "yes"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row: list[Any], idx_map: dict[str, int], key: str, default: Any = "") -> Any:
    idx = idx_map.get(key)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def import_asphalt_rates(db: Session, content: bytes, filename: str) -> dict[str, Any]:
    table = parse_tabular(content, filename)
    if not table:
        raise ValueError("File is empty")
    headers = [str(h or "") for h in table[0]]
    idx = _header_map(headers)
    if "name" not in idx or "subcontractor" not in idx:
        raise ValueError("Need at least subcontractor and name columns")

    created = updated = skipped = 0
    errors: list[str] = []
    created_subs = 0
    for i, raw in enumerate(table[1:], start=2):
        sub_name = str(_cell(raw, idx, "subcontractor") or "").strip()
        name = str(_cell(raw, idx, "name") or "").strip()
        if not sub_name or not name:
            skipped += 1
            continue
        try:
            sub = (
                db.query(AsphaltSubcontractor)
                .filter(func.lower(AsphaltSubcontractor.name) == sub_name.lower())
                .first()
            )
            if not sub:
                max_pos = db.query(func.max(AsphaltSubcontractor.position)).scalar() or 0
                sub = AsphaltSubcontractor(name=sub_name, position=max_pos + 10)
                db.add(sub)
                db.flush()
                created_subs += 1
            unit = normalize_unit(str(_cell(raw, idx, "unit") or "m2"))
            rate_type = str(_cell(raw, idx, "rate_type") or "").strip().lower() or infer_rate_type(unit, name)
            if rate_type not in ("unit", "shift"):
                rate_type = infer_rate_type(unit, name)
            unit_rate = _as_float(_cell(raw, idx, "unit_rate"))
            day = _as_float(_cell(raw, idx, "day_rate"), unit_rate)
            night = _as_float(_cell(raw, idx, "night_rate"))
            weekend = _as_float(_cell(raw, idx, "weekend_rate")) or _as_float(_cell(raw, idx, "saturday_rate"))
            ph = _as_float(_cell(raw, idx, "public_holiday_rate"))
            stored = apply_stored_rates(
                {
                    "day_rate": day or unit_rate,
                    "night_rate": night,
                    "sunday_rate": weekend,
                    "saturday_rate": weekend,
                    "public_holiday_rate": ph,
                    "unit_rate": unit_rate or None,
                },
                rate_type=rate_type,
            )
            existing = (
                db.query(AsphaltRate)
                .filter(
                    AsphaltRate.subcontractor_id == sub.id,
                    func.lower(AsphaltRate.name) == name.lower(),
                )
                .first()
            )
            if existing:
                existing.unit = unit
                existing.rate_type = rate_type
                existing.day_rate = stored["day_rate"]
                existing.night_rate = stored["night_rate"]
                existing.saturday_rate = stored["saturday_rate"]
                existing.sunday_rate = stored["sunday_rate"]
                existing.public_holiday_rate = stored["public_holiday_rate"]
                existing.active = True
                updated += 1
            else:
                max_rpos = db.query(func.max(AsphaltRate.position)).scalar() or 0
                db.add(
                    AsphaltRate(
                        subcontractor_id=sub.id,
                        name=name,
                        unit=unit,
                        rate_type=rate_type,
                        day_rate=stored["day_rate"],
                        night_rate=stored["night_rate"],
                        saturday_rate=stored["saturday_rate"],
                        sunday_rate=stored["sunday_rate"],
                        public_holiday_rate=stored["public_holiday_rate"],
                        position=max_rpos + 1,
                    )
                )
                created += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Row {i} ({sub_name} / {name}): {exc}")
    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "subcontractors_created": created_subs,
        "errors": errors[:50],
    }


def import_traffic_rates(db: Session, content: bytes, filename: str) -> dict[str, Any]:
    table = parse_tabular(content, filename)
    if not table:
        raise ValueError("File is empty")
    idx = _header_map([str(h or "") for h in table[0]])
    if "name" not in idx:
        raise ValueError("Need a name column")
    created = updated = skipped = 0
    errors: list[str] = []
    for i, raw in enumerate(table[1:], start=2):
        name = str(_cell(raw, idx, "name") or "").strip()
        if not name:
            skipped += 1
            continue
        try:
            weekend_o = _as_float(_cell(raw, idx, "weekend_ordinary"))
            weekend_t = _as_float(_cell(raw, idx, "weekend_overtime"))
            payload = {
                "name": name,
                "rate_kind": str(_cell(raw, idx, "kind") or "crew_pack").strip() or "crew_pack",
                "pack_people": int(_as_float(_cell(raw, idx, "pack_people"), 1)),
                "includes_vehicle": _as_bool(_cell(raw, idx, "includes_vehicle"), False),
                "day_ordinary": _as_float(_cell(raw, idx, "day_ordinary")),
                "day_overtime": _as_float(_cell(raw, idx, "day_overtime")),
                "night_ordinary": _as_float(_cell(raw, idx, "night_ordinary")),
                "night_overtime": _as_float(_cell(raw, idx, "night_overtime")),
                "saturday_ordinary": weekend_o,
                "saturday_overtime": weekend_t,
                "sunday_ordinary": weekend_o,
                "sunday_overtime": weekend_t,
                "public_holiday_ordinary": _as_float(_cell(raw, idx, "public_holiday_ordinary")),
                "public_holiday_overtime": _as_float(_cell(raw, idx, "public_holiday_overtime")),
                "active": _as_bool(_cell(raw, idx, "active"), True),
            }
            existing = db.query(LabourRate).filter(func.lower(LabourRate.name) == name.lower()).first()
            if existing:
                for k, v in payload.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                max_pos = db.query(func.max(LabourRate.position)).scalar() or 0
                db.add(LabourRate(position=max_pos + 1, **payload))
                created += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Row {i} ({name}): {exc}")
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:50]}
