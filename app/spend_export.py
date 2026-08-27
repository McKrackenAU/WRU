"""Excel and PDF exports for actual traffic / asphalt spend."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .pdf_brand import GREEN, GREEN_HEX, MUTED, ROW_ALT, RULE, branded_margins, draw_branded_page


def _money(n: Any) -> str:
    try:
        return f"${float(n or 0):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def _kind_label(kind: str | None) -> str:
    k = (kind or "").lower()
    if k == "asphalt":
        return "Pavements / asphalt"
    if k == "traffic":
        return "Traffic"
    return kind or "—"


def _fmt_date(value: Any) -> str:
    if not value:
        return "—"
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    text = str(value)[:10]
    try:
        y, m, d = text.split("-")
        return f"{d}/{m}/{y}"
    except ValueError:
        return text


def build_spend_workbook(rows: list[dict[str, Any]], *, title: str = "Actual spend") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Spend"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=GREEN_HEX)
    money_font = Font(bold=True)

    ws.cell(1, 1, "WRU TGS Tracker — Actual spend").font = Font(bold=True, size=14)
    ws.cell(2, 1, title)
    ws.cell(3, 1, f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(4, 1, f"{len(rows)} row(s)")

    headers = [
        "Kind",
        "Date",
        "Road",
        "Site",
        "Program",
        "Contractor",
        "Category",
        "Source",
        "Invoice",
        "Amount",
        "Notes",
        "Entered by",
    ]
    start = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(start, col, h)
        cell.font = header_font
        cell.fill = header_fill

    total = 0.0
    for i, row in enumerate(rows):
        r = start + 1 + i
        contractor = row.get("contractor_name") or "—"
        amount = float(row.get("amount") or 0)
        total += amount
        src = row.get("source") or "manual"
        if src == "calculated":
            source = "From rates"
        elif src == "from_estimate":
            source = "From estimate"
        else:
            source = "Manual"
        values = [
            _kind_label(row.get("kind")),
            _fmt_date(row.get("work_date")),
            row.get("road_name") or "—",
            row.get("site_number") or "—",
            row.get("program") or "—",
            contractor,
            row.get("category") or "—",
            source,
            row.get("invoice_ref") or "—",
            amount,
            row.get("notes") or "",
            row.get("created_by") or "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(r, col, val)
            if col == 10:
                cell.number_format = '"$"#,##0.00'
                cell.font = money_font

    total_row = start + 1 + len(rows)
    ws.cell(total_row, 9, "Total").font = Font(bold=True)
    total_cell = ws.cell(total_row, 10, total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'

    from openpyxl.utils import get_column_letter

    widths = [16, 12, 28, 12, 20, 22, 14, 12, 16, 12, 32, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_spend_pdf(rows: list[dict[str, Any]], *, title: str = "Actual spend") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title=title,
        **branded_margins(landscape_mode=True),
    )
    doc.brand_eyebrow = "ACTUAL SPEND"
    doc.brand_title = title
    total = sum(float(r.get("amount") or 0) for r in rows)
    doc.brand_subtitle = f"{len(rows)} row(s) · total {_money(total)}"
    doc.brand_doc_kind = "Spend"
    doc.brand_product = "WRU TGS TRACKER"
    doc.brand_footer_meta = f"{len(rows)} row(s)"

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="Th",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.5,
            leading=9,
            textColor=colors.white,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Td",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#1a1a1a"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="FootNote",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=MUTED,
            spaceBefore=6,
        )
    )

    headers = ["Kind", "Date", "Road", "Site", "Contractor", "Source", "Invoice", "Amount", "Notes"]
    data: list[list[Any]] = [[Paragraph(h, styles["Th"]) for h in headers]]
    if not rows:
        data.append([Paragraph("No spend rows match these filters.", styles["Td"])] + [""] * 8)
    else:
        for row in rows:
            source = row.get("source") or "manual"
            if source == "calculated":
                source_label = "From rates"
            elif source == "from_estimate":
                source_label = "From estimate"
            else:
                source_label = "Manual"
            data.append(
                [
                    Paragraph(_kind_label(row.get("kind")), styles["Td"]),
                    Paragraph(_fmt_date(row.get("work_date")), styles["Td"]),
                    Paragraph(str(row.get("road_name") or "—")[:36], styles["Td"]),
                    Paragraph(str(row.get("site_number") or "—"), styles["Td"]),
                    Paragraph(str(row.get("contractor_name") or "—")[:28], styles["Td"]),
                    Paragraph(source_label, styles["Td"]),
                    Paragraph(str(row.get("invoice_ref") or "—")[:18], styles["Td"]),
                    Paragraph(_money(row.get("amount")), styles["Td"]),
                    Paragraph(str(row.get("notes") or "")[:80], styles["Td"]),
                ]
            )

    col_w = [28 * mm, 20 * mm, 42 * mm, 18 * mm, 36 * mm, 22 * mm, 24 * mm, 22 * mm, 50 * mm]
    table = Table(data, colWidths=col_w, repeatRows=1)
    cmds: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), GREEN),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, RULE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("ALIGN", (7, 1), (7, -1), "RIGHT"),
    ]
    if not rows:
        cmds.append(("SPAN", (0, 1), (-1, 1)))
    table.setStyle(TableStyle(cmds))

    story = [
        table,
        Spacer(1, 4 * mm),
        Paragraph(f"Total {_money(total)} · Ventia confidential.", styles["FootNote"]),
    ]
    doc.build(story, onFirstPage=draw_branded_page, onLaterPages=draw_branded_page)
    return buf.getvalue()
