"""Build Excel and PDF exports for comms planner tabs."""

from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .pdf_brand import GREEN, GREEN_HEX, ROW_ALT, RULE, branded_margins

SPECIAL_COLUMNS = (
    ("_job", "Job"),
    ("_linked_site", "Linked job"),
    ("_files", "Files"),
)

PDF_CHUNK = 7


def _job_label(row: dict) -> str:
    site = row.get("site") or {}
    road = (site.get("road_name") or "").strip()
    number = (site.get("site_number") or "").strip()
    if road and number:
        return f"{road} · {number}"
    return road or ""


def _cell_value(row: dict, key: str) -> str:
    if key == "_job":
        values = row.get("values") or {}
        return (
            str(values.get("location") or values.get("road_street_name") or "").strip()
            or _job_label(row)
            or row.get("section")
            or ""
        )
    if key == "_linked_site":
        return _job_label(row)
    if key == "_files":
        count = row.get("document_count")
        return "" if count in (None, "") else str(count)
    raw = (row.get("values") or {}).get(key)
    if raw is None:
        return ""
    return str(raw)


def _headers(columns: list[dict], keys: list[str], include_job: bool) -> list[tuple[str, str]]:
    by_key = {col["field_key"]: col["name"] for col in columns}
    out: list[tuple[str, str]] = []
    if include_job and "_job" not in keys:
        out.append(("_job", "Job"))
    for key in keys:
        if key in by_key:
            out.append((key, by_key[key]))
        else:
            special = next((name for sk, name in SPECIAL_COLUMNS if sk == key), None)
            if special:
                out.append((key, special))
    return out


def collect_export_tables(
    sheets: list[dict],
    *,
    column_keys: list[str],
    row_ids: list[int] | None,
    include_job: bool,
) -> list[dict[str, Any]]:
    wanted = set(row_ids or [])
    tables = []
    for sheet in sheets:
        columns = sheet.get("columns") or []
        headers = _headers(columns, column_keys, include_job)
        rows = []
        for row in sheet.get("rows") or []:
            if wanted and row.get("id") not in wanted:
                continue
            rows.append([_cell_value(row, key) for key, _name in headers])
        tables.append(
            {
                "title": sheet.get("title") or "Planner",
                "headers": [name for _key, name in headers],
                "rows": rows,
            }
        )
    return tables


def build_comms_xlsx(tables: list[dict[str, Any]], *, title: str) -> bytes:
    wb = Workbook()
    first = True
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=GREEN_HEX)
    for table in tables:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = str(table["title"])[:31] or "Planner"
        ws.cell(1, 1, title).font = Font(bold=True, size=13)
        ws.cell(2, 1, f"Exported {date.today().isoformat()} · {len(table['rows'])} row(s)")
        for col, name in enumerate(table["headers"], 1):
            cell = ws.cell(4, col, name)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, row in enumerate(table["rows"], start=5):
            for c_idx, value in enumerate(row, 1):
                ws.cell(r_idx, c_idx, value)
        if not table["headers"]:
            ws.cell(4, 1, "No columns selected")
        for col in ws.columns:
            letter = col[0].column_letter
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[letter].width = min(max(width + 2, 12), 48)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_cell(text: Any, style: ParagraphStyle) -> Paragraph:
    raw = "" if text is None else str(text)
    safe = (
        raw.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(safe or "—", style)


def build_comms_pdf(tables: list[dict[str, Any]], *, title: str) -> bytes:
    from .pdf_brand import draw_branded_page

    buf = io.BytesIO()
    total = sum(len(table["rows"]) for table in tables)
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title=title,
        **branded_margins(landscape_mode=True),
    )
    caption = f"{len(tables)} tab(s) · {total} row(s)"
    doc.brand_eyebrow = "COMMS PLANNER"
    doc.brand_title = title
    doc.brand_subtitle = caption
    doc.brand_doc_kind = "Comms Planner"
    doc.brand_product = "WRU TGS TRACKER"
    doc.brand_footer_meta = caption

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CommsTh",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            leading=9,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CommsTd",
            parent=styles["Normal"],
            fontSize=7,
            leading=9,
            textColor=colors.HexColor("#1a1a1a"),
            alignment=TA_LEFT,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CommsH",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=GREEN,
            spaceBefore=8,
            spaceAfter=4,
        )
    )

    usable = 273 * mm
    story: list[Any] = []
    for table in tables:
        headers = table["headers"] or ["No columns selected"]
        rows = table["rows"] or [[]]
        chunks = [headers[i : i + PDF_CHUNK] for i in range(0, max(len(headers), 1), PDF_CHUNK)] or [headers]
        story.append(Paragraph(str(table["title"]), styles["CommsH"]))
        for chunk_index, chunk in enumerate(chunks):
            start = chunk_index * PDF_CHUNK
            width = usable / max(len(chunk), 1)
            data = [[_pdf_cell(name, styles["CommsTh"]) for name in chunk]]
            if not table["rows"]:
                data.append([_pdf_cell("No rows selected.", styles["CommsTd"])] + [""] * (len(chunk) - 1))
            else:
                for row in table["rows"]:
                    slice_row = row[start : start + len(chunk)]
                    while len(slice_row) < len(chunk):
                        slice_row.append("")
                    data.append([_pdf_cell(value, styles["CommsTd"]) for value in slice_row])
            grid = Table(data, colWidths=[width] * len(chunk), repeatRows=1)
            cmds: list[tuple] = [
                ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("GRID", (0, 0), (-1, -1), 0.25, RULE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
            ]
            if not table["rows"]:
                cmds.append(("SPAN", (0, 1), (-1, 1)))
            grid.setStyle(TableStyle(cmds))
            story.append(grid)
            story.append(Spacer(1, 6))

    doc.build(story, onFirstPage=draw_branded_page, onLaterPages=draw_branded_page)
    return buf.getvalue()
