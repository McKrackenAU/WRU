"""Tracker spreadsheet import — status aliases and unmatched rows."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from app.tracker_import import _status_workflow, parse_tracker_workbook


def test_status_aliases_ready_for_works():
    wf, unmatched = _status_workflow("Ready For Works")
    assert unmatched is None
    assert wf["ready_for_works"] is True
    assert wf["moa_received"] is True
    assert wf["ventia_review"] is True


def test_moa_received_matches_ready_for_works():
    wf, unmatched = _status_workflow("MoA received / approved")
    assert unmatched is None
    assert wf["moa_received"] is True
    assert wf["ready_for_works"] is True


def test_status_submitted_to_traffic_management():
    wf, unmatched = _status_workflow("Submitted to traffic management")
    assert unmatched is None
    assert wf["submitted_to_tmd"] is True
    assert wf["ready_for_works"] is False


def test_blank_and_no_import_as_not_started():
    wf, unmatched = _status_workflow("")
    assert unmatched is None
    assert wf["tgs_markup_completed"] is False
    wf2, _ = _status_workflow("No")
    assert wf2["ready_for_works"] is False


def test_unknown_status_does_not_drop_row():
    wf, unmatched = _status_workflow("Waiting on designer")
    assert unmatched == "Waiting on designer"
    assert wf["ready_for_works"] is False


def test_short_status_does_not_match_longer_alias():
    wf, unmatched = _status_workflow("Received")
    assert unmatched == "Received"
    assert wf["plan_received"] is False
    assert wf["moa_received"] is False


def test_status_with_comment_suffix_still_maps():
    wf, unmatched = _status_workflow("MoA Submitted - TRIMS requested 22/7/26")
    assert unmatched is None
    assert wf["moa_submitted"] is True
    assert wf["moa_with_trims"] is False


def test_yes_uses_lamps_instead_of_jumping_to_ready_for_works():
    lamps = [None, None, None, None, 0, 0, 0, 0]  # through MoA Submitted
    wf, unmatched = _status_workflow("Yes", lamps)
    assert unmatched is None
    assert wf["moa_submitted"] is True
    assert wf["ready_for_works"] is False


def test_all_zero_lamps_stay_not_started():
    wf, unmatched = _status_workflow("", [0, 0, 0, 0, 0, 0, 0, 0])
    assert unmatched is None
    assert wf["tgs_markup_completed"] is False
    assert wf["submitted_to_tmd"] is False


def _tracker_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TGS-MOA Tracker"
    ws["A1"] = "LCP - FMRP"
    ws["B1"] = "Road Name"
    ws["B3"] = "HIGH ST - 1234"
    ws["C3"] = "S48"
    ws["G3"] = "Ready For Works"
    ws["R3"] = "0093225"
    ws["B4"] = "LOW ST - 99"
    ws["C4"] = "S1"
    ws["G4"] = "Waiting on designer"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_keeps_unmatched_status_rows():
    parsed = parse_tracker_workbook(_tracker_bytes())
    assert parsed["parsed"] == 2
    roads = {r["road_name"] for r in parsed["rows"]}
    assert "HIGH ST - 1234" in roads
    assert "LOW ST - 99" in roads
    unmatched_row = next(r for r in parsed["rows"] if r["site_number"] == "S1")
    assert unmatched_row["status_unmatched"]
    assert "Waiting on designer" in parsed["unmatched_statuses"]


def _v6_like_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TGS-MOA Tracker"
    headers = {
        2: "Road Name",
        3: "Site Number",
        4: "Indicative Site Start Date",
        7: "TGS Markup completed",
        16: "Comments",
        18: "MoA Number",
        19: "MoA Submission Date",
        24: "Council",
        25: "Council Submission Date",
        26: "Council No Objection Recieved Date",
        35: "JOB COMPLETED",
    }
    for col, label in headers.items():
        ws.cell(1, col, label)
    ws["A2"] = "LCP - FMRP"

    ws["A3"] = 1
    ws["B3"] = "DYNON RD - 5035"
    ws["C3"] = "S48"
    ws["D3"] = datetime(2026, 9, 13)
    ws["G3"] = "MoA Submitted"
    ws["P3"] = "TRIMS requested 22/7/26\nVentia 28/7 responded to DTP"
    ws["R3"] = "0093225"
    ws["S3"] = datetime(2026, 7, 3)
    ws["X3"] = "Maribyrnong"
    ws["Y3"] = datetime(2026, 7, 6)
    ws["Z3"] = datetime(2026, 7, 28)
    ws["AI3"] = "No"

    ws["B65"] = "ADD NEW LINE ABOVE"
    ws["B66"] = "TOTALS"
    ws["C66"] = "COMPLETED"
    ws["G66"] = 62
    ws["B68"] = 62
    ws["C68"] = "% COMPLETE"
    ws["G68"] = 1

    ws["A87"] = "GENERICS MTMP & ITMP"
    ws["A88"] = 1
    ws["B88"] = "WRU - Miepol-TMR-VEN-OTH-396-MTMP 25-26"
    ws["G88"] = "Yes"
    ws["P88"] = "TGS-TMR-VEN-OTH-396\nChange form V2"
    ws["R88"] = 77388
    ws["X88"] = "Multiple"
    ws["Y88"] = datetime(2025, 7, 15)
    ws["Z88"] = "Brim-30/07/25\nMel-17/07/25\nHob-"
    ws["AI88"] = "No"

    ws["A81"] = "STRUCTURES"
    ws["A82"] = 10
    ws["B82"] = "PRINCES HWY WEST BRIDGE"
    ws["G82"] = "Yes"
    ws["P82"] = "TGS-TMR-GRD-2500-374\nsubmitted on 24-Feb-25"
    ws["X82"] = "Multiple"
    ws["Y82"] = datetime(2024, 9, 10)
    ws["AI82"] = "Yes"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_skips_totals_and_keeps_comments_and_council_dates():
    parsed = parse_tracker_workbook(_v6_like_bytes())
    roads = {r["road_name"] for r in parsed["rows"]}
    assert "DYNON RD - 5035" in roads
    assert "% COMPLETE" not in roads
    assert "62" not in roads
    assert parsed["parsed"] == 3

    dynon = next(r for r in parsed["rows"] if r["site_number"] == "S48")
    assert "TRIMS requested 22/7/26" in dynon["comments"]
    assert "\n" in dynon["comments"]
    assert dynon["councils"] == [
        {
            "council_name": "Maribyrnong",
            "submitted_to_council_date": date(2026, 7, 6),
            "no_objection_date": date(2026, 7, 28),
        }
    ]
    assert dynon["must_have_manual"] is False
    assert dynon["include_in_totals"] is True

    generic = next(r for r in parsed["rows"] if "Miepol" in r["road_name"])
    names = {c["council_name"]: c for c in generic["councils"]}
    assert set(names) == {"Brimbank", "Melbourne", "Hobsons Bay"}
    assert names["Brimbank"]["no_objection_date"].isoformat() == "2025-07-30"
    assert names["Melbourne"]["no_objection_date"].isoformat() == "2025-07-17"
    assert names["Brimbank"]["submitted_to_council_date"].isoformat() == "2025-07-15"
    assert generic["is_generic_moa"] is True
    assert generic["tgs_reference"] == "TGS-TMR-VEN-OTH-396"
    assert generic["moa_number"] == "77388"

    structures = next(r for r in parsed["rows"] if "BRIDGE" in r["road_name"])
    assert structures["include_in_totals"] is False
    assert structures["tgs_reference"] == "TGS-TMR-GRD-2500-374"
    assert structures["archive"] is True
    assert "1" not in parsed["unmatched_statuses"]


def test_real_v6_workbook_if_attached():
    path = Path("/home/ubuntu/.cursor/projects/workspace/uploads/WRU_Traffic_TGS-MOA_Tracker_V6_WIP_f68a.xlsm")
    if not path.is_file():
        return
    parsed = parse_tracker_workbook(path.read_bytes())
    roads = {r["road_name"] for r in parsed["rows"]}
    assert "DYNON RD - 5035" in roads
    assert not any("%" in r["site_number"] for r in parsed["rows"])
    dynon = next(r for r in parsed["rows"] if r["site_number"] == "S48")
    assert "TRIMS requested" in (dynon["comments"] or "")
    assert dynon["workflow"]["moa_submitted"] is True
    assert dynon["workflow"]["ready_for_works"] is False
    last_stages: dict[str, int] = {}
    for row in parsed["rows"]:
        last = "none"
        for k in (
            "tgs_markup_completed",
            "submitted_to_tmd",
            "plan_received",
            "ready_to_submit_moa",
            "moa_submitted",
            "moa_with_trims",
            "moa_received",
            "ready_for_works",
        ):
            if row["workflow"].get(k):
                last = k
        last_stages[last] = last_stages.get(last, 0) + 1
    assert last_stages.get("submitted_to_tmd", 0) >= 30
    assert last_stages.get("moa_submitted", 0) >= 10
    assert dynon["councils"][0]["council_name"] == "Maribyrnong"
    assert str(dynon["councils"][0]["submitted_to_council_date"]) == "2026-07-06"
    assert str(dynon["councils"][0]["no_objection_date"]) == "2026-07-28"
    miepol = next(r for r in parsed["rows"] if "Miepol" in r["road_name"])
    names = {c["council_name"] for c in miepol["councils"]}
    assert "Brimbank" in names
    assert "Melbourne" in names
    assert "Hobsons Bay" in names
    brim = next(c for c in miepol["councils"] if c["council_name"] == "Brimbank")
    assert str(brim["no_objection_date"]) == "2025-07-30"
    assert parsed["parsed"] >= 60
    assert parsed["parsed"] <= 72
