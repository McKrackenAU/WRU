"""Priority list export respects the lists page Show-top cap and filters."""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.routers.export import (
    CLIENT_FIELDS,
    CLIENT_HEADERS,
    ClientListView,
    _client_export_values,
    _pdf_bytes,
    _xlsx_bytes,
    apply_client_list_view,
)

ROOT = Path(__file__).resolve().parent.parent
LISTS_JS = (ROOT / "app/static/js/lists.js").read_text(encoding="utf-8")
LISTS_HTML = (ROOT / "app/static/lists.html").read_text(encoding="utf-8")
EXPORT = (ROOT / "app/routers/export.py").read_text(encoding="utf-8")
INDEX_HTML = (ROOT / "app/static/index.html").read_text(encoding="utf-8")
APP_JS = (ROOT / "app/static/js/app.js").read_text(encoding="utf-8")


def _row(road, *, pri=1, rank=1, program="Assets", start="2026-10-01", wait=3, moa="M1"):
    return {
        "priority": pri,
        "rank": rank,
        "road_name": road,
        "site_number": road[:1],
        "program": program,
        "indicative_start": start,
        "business_days_waiting": wait,
        "moa_number": moa,
    }


def test_limit_takes_the_first_n_after_sort():
    rows = [
        _row("C Road", start="2026-12-01", rank=3),
        _row("A Road", start="2026-10-01", rank=1),
        _row("B Road", start="2026-11-01", rank=2),
    ]
    out = apply_client_list_view(
        rows,
        ClientListView(limit=2, sort="start", direction="asc"),
    )
    assert [r["road_name"] for r in out] == ["A Road", "B Road"]


def test_limit_none_returns_all():
    rows = [_row("A"), _row("B"), _row("C")]
    out = apply_client_list_view(rows, ClientListView(limit=None))
    assert len(out) == 3


def test_empty_priority_filter_exports_nothing():
    rows = [_row("A", pri=1), _row("B", pri=2)]
    out = apply_client_list_view(rows, ClientListView(priorities=set(), limit=20))
    assert out == []


def test_program_and_priority_filters():
    rows = [
        _row("Keep", pri=1, program="Assets"),
        _row("Skip pri", pri=2, program="Assets"),
        _row("Skip prog", pri=1, program="Lifecycle"),
    ]
    out = apply_client_list_view(
        rows,
        ClientListView(priorities={"1"}, programs={"Assets"}),
    )
    assert [r["road_name"] for r in out] == ["Keep"]


def test_lists_page_wires_export_query_to_cap():
    assert "syncExportLinks" in LISTS_JS
    assert 'params.set("limit", String(state.cap))' in LISTS_JS
    assert "js-list-export" in LISTS_HTML
    assert 'data-export="/api/export/permits-list.pdf"' in LISTS_HTML
    assert 'data-export="/api/export/trims-list.pdf"' in LISTS_HTML
    assert "Export top" in LISTS_JS
    assert "client_list_view" in EXPORT
    assert "apply_client_list_view" in EXPORT
    assert "limit: int | None = Query" in EXPORT


def test_client_export_is_six_columns_in_requested_order():
    assert CLIENT_FIELDS == [
        "priority",
        "moa_number",
        "road_name",
        "site_number",
        "program",
        "indicative_start",
    ]
    assert CLIENT_HEADERS == [
        "Priority",
        "MoA Number",
        "Road",
        "Site",
        "Program",
        "Indicative start",
    ]
    values = _client_export_values(_row("FOOTSCRAY-SUNSHINE RD - 5877", moa="MOA-0094645"))
    assert list(values) == CLIENT_HEADERS
    assert values["MoA Number"] == "MOA-0094645"
    assert "Must-have" not in values
    assert "Stage" not in EXPORT.split("headers = [", 1)[1].split("]", 1)[0]


def test_pdf_and_xlsx_use_slim_columns():
    rows = [_row("FOOTSCRAY-SUNSHINE RD - 5877", moa="MOA-0094645", program="Structures")]
    header_block = EXPORT.split("headers = [", 1)[1].split("]", 1)[0]
    assert '"Pri"' in header_block
    assert '"MoA #"' in header_block
    assert '"Road"' in header_block
    assert '"Site"' in header_block
    assert '"Program"' in header_block
    assert '"Indicative start"' in header_block
    assert "Must-have" not in header_block
    assert "Council / wait" not in header_block
    assert "Comments" not in header_block
    # PDF streams are compressed; generating still proves the 6-column table builds.
    pdf = _pdf_bytes(rows, title="DTP — Permits priority list", team_label="Permits team")
    assert pdf.startswith(b"%PDF")
    assert b"DTP" in pdf

    wb = load_workbook(BytesIO(_xlsx_bytes(rows, "Permits", "Permits list")))
    ws = wb.active
    headers = [ws.cell(4, col).value for col in range(1, 8)]
    assert headers[:6] == CLIENT_HEADERS
    assert headers[6] is None
    assert ws.cell(5, 2).value == "MOA-0094645"
    assert ws.cell(5, 3).value == "FOOTSCRAY-SUNSHINE RD - 5877"
    assert ws.cell(5, 6).value == "2026-10-01"


def test_road_name_is_list_select_with_other():
    assert 'id="fRoadSelect"' in INDEX_HTML
    assert 'id="fRoadOther"' in INDEX_HTML
    assert 'id="roadList"' not in INDEX_HTML
    assert 'list="roadList"' not in INDEX_HTML
    assert "ROAD_OTHER" in APP_JS
    assert "Other…" in APP_JS
    assert "collectedRoadName" in APP_JS
    assert "used_roads" in (ROOT / "app/main.py").read_text(encoding="utf-8")
